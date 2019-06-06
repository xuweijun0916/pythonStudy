import openpyxl,os
os.chdir(r"C:\Users\xuwj\Desktop\Automate the Boring Stuff with Python书学习资料\automate_online-materials")
#os.getcwd()
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name("Sheet")

price_update = {
    'Garlic':3.07,
    'Celery':1.19,
    'Lemon':1.27
}
for rowNum in range(2,sheet.max_row):
    productName = sheet.cell(row=rowNum,column=1).value
    if productName in price_update:
        sheet.cell(row=rowNum,column=2).value = price_update[productName]
wb.save('updatedProduceSales.xlsx')