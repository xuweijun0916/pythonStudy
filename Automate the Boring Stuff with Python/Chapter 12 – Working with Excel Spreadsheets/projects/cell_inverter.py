import openpyxl

wb = openpyxl.load_workbook('before.xlsx')
sheet = wb.active

nwb = openpyxl.Workbook()
nst = nwb.active

"""把一个表格中的数据全部导出到一个列表"""
listResult = []
for i in range(1, sheet.max_row + 1):
    lineData = []
    for j in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=i, column=j)
        lineData.append(cell.value)
    listResult.append(lineData)

for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        nst.cell(row=j, column=i).value = listResult[i-1][j-1]

print(listResult)
nwb.save('after.xlsx')