import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active

nwb = openpyxl.Workbook()
nst = wb.active


def balckInsert(n, line):
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            if row < n:
                nst.cell(row=row, column=column).value = sheet.cell(row=row, column=column).value
            else:
                nst.cell(row=row + line, column=column).value = sheet.cell(row=row, column=column).value

    nwb.save("nwb.xlsx")


if __name__ == "__main__":
    balckInsert(3, 5)
