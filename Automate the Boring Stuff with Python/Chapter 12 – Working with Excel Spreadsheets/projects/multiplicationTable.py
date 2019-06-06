import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active
bold = Font(bold=True)
for i in range(1, 7):
    sheet.cell(row=i + 1, column=1).value = i
    sheet.cell(row=i + 1, column=1).font = bold
    sheet.cell(row=1, column=i + 1).value = i
    sheet.cell(row=1, column=i + 1).font = bold

for i in range(2, 8):
    for j in range(2, 8):
        x = sheet.cell(row=i, column=1).value
        y = sheet.cell(row=1, column=j).value
        sheet.cell(row=i, column=j).value = x * y

wb.save("multiplicationTable.xlsx")